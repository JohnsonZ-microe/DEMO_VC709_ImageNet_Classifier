import openpyxl


def extract_data_from_same_cell(file_path, cell_address, output_file="extracted_data.xlsx"):
    """
    提取xlsx文件中所有工作表同一位置的数据

    参数:
        file_path: 输入的xlsx文件路径
        cell_address: 要提取的单元格地址 (例如 "C5")
        output_file: 结果保存的文件名 (可选)
    """

    # 加载工作簿 (data_only=True 表示读取公式计算后的结果，而非公式本身)
    wb = openpyxl.load_workbook(file_path, data_only=True)

    # 用于存储结果
    results = []

    # 遍历所有工作表
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        # 获取指定单元格的值
        cell_value = ws[cell_address].value
        results.append({
            "工作表名称": sheet_name,
            f"单元格 {cell_address} 的值": cell_value
        })
        print(f"工作表: {sheet_name} -> {cell_address} = {cell_value}")

    # --- 将结果保存到新的Excel文件中 ---
    wb_out = openpyxl.Workbook()
    ws_out = wb_out.active
    ws_out.title = "提取结果"

    # 写入表头
    ws_out.append(["工作表名称", f"单元格 {cell_address} 的值"])

    # 写入数据
    for row in results:
        ws_out.append([row["工作表名称"], row[f"单元格 {cell_address} 的值"]])

    # 保存
    wb_out.save(output_file)
    print(f"\n数据已提取并保存至: {output_file}")


# ================= 配置区域 =================
# 在这里修改你的文件名和单元格位置
INPUT_FILE = "ofmp11.xlsx"  # 改成你的xlsx文件路径
TARGET_CELL = "K18"  # 改成你想提取的单元格位置，例如 "A1", "D10"

# ================= 运行程序 =================
if __name__ == "__main__":
    extract_data_from_same_cell(INPUT_FILE, TARGET_CELL)